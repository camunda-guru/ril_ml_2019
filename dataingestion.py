# -*- coding: utf-8 -*-
"""
Created on Thu Nov 21 11:20:43 2019

@author: Balasubramaniam
"""

import pymysql;
from openpyxl import load_workbook

filePath="users.xlsx"

def createConnection():
    
    conn=pymysql.connect(host="localhost",user="root",
                   passwd="vignesh",
                   db="jiodb");
    return conn;



def readExcelData():
    fileRef=load_workbook(filePath)
    sheetRef=fileRef.get_sheet_by_name("user-data")
    outerList=[]
    for row in range(1,10):
        innerList=[]
        for col in range(1,5):
            print(sheetRef.cell(row=row,column=col).value) 
            innerList.append(sheetRef.cell(row=row,column=col).value) 
        outerList.append(innerList)    
    return outerList         
            

def insertCustomerData():
    conn=createConnection()
    data=readExcelData()    
    cursor=conn.cursor()
    try:
        for innerdata in data:
            cursor.execute("""insert into customer values 
                           ('%d','%s','%s','%s')""" 
                           %(int(innerdata[0]),innerdata[1],innerdata[2],innerdata[3]));
            #cursor.execute(query);
            conn.commit()
    except pymysql.Error as e:
        print("Exception occurred",e )
        conn.rollback()

    finally:
        conn.close()



insertCustomerData()

#print(readExcelData())








'''

def fetchBytestCaseId(id):
    conn=createConnection()
    cursor=conn.cursor()
    cursor.execute("""select * from testcaseresults where 
    tcno='%d'""" %(id))
    return cursor.fetchall()


def addTestCaseInfo(data):
    connObj=createConnection();
    cursor=connObj.cursor();
    print("Cursor ready...");     
    try:
        cursor.execute("""insert into testcaseresults values 
    ('%d','%s','%d')""" %(data[0],data[1],data[2]));
        #cursor.execute(query);
        connObj.commit()
    except pymysql.Error as e:
        print("Exception occurred",e )
        connObj.rollback()

    finally:
        connObj.close()
    
#tsData=[1099,'REQ1099',1];
#addTestCaseInfo(tsData);
rows = fetchBytestCaseId(1002);

for row in rows:
    print(row);

'''

