# -*- coding: utf-8 -*-
"""
Created on Mon Aug 21 14:05:06 2017

@author: BALASUBRAMANIAM
"""
import requests
from openpyxl import Workbook,load_workbook
response= requests.get("http://jsonplaceholder.typicode.com/users")
print(response)

#for user in response.json():
    #print(user['name'],'\t',user['email'],'\t',user['address']['city'])

#print( response.json()[0])
filePath="F:\\citi_ml_jun2018\\day6\\users.xlsx"
'''
wb=Workbook()
wb.create_sheet("user-data")
wb.save(filePath)
'''
fileRef=load_workbook(filePath,read_only=False)
sheetRef=fileRef.get_sheet_by_name("user-data")
length=len(response.json())

for row in range(1,length):
    col=0
    for (key,value) in response.json()[row].items():
        print(key,value)
        col=col+1
        if(col<5):
         sheetRef.cell(column=col,row=row,value="%s" % value)


fileRef.save(filePath)


'''
finData=requests.get("https://www.quandl.com/api/v3/datasets/OPEC/ORB.json",verify=False)
print(finData)

print(finData.json()['dataset'])

directionsData=requests.get("https://maps.googleapis.com/maps/api/directions/json?origin=Bangalore&destination=Pune",verify=False)
print(directionsData.json()['routes'][0]['bounds'])
'''





